from telegram import Update, ReplyKeyboardMarkup
from telegram.constants import ParseMode
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
from datetime import datetime, timedelta, time as datetime_time
import json
import os
from zoneinfo import ZoneInfo

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    Workbook = None

BANGLADESH_TZ = ZoneInfo("Asia/Dhaka")

# ---------------- CONFIG ----------------
TOKEN = os.environ.get("TOKEN")  # Replit Secret
ADMIN_CHAT_ID = 7898548948
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_FILE = os.path.join(BASE_DIR, "tracker_data.json")
REPORTS_DIR = os.path.join(BASE_DIR, "reports")

SMOKE_LIMIT_COUNT = 5
TOILET_LIMIT_COUNT = 5
EAT_LIMIT_COUNT = 1
SMOKE_WARNING_MINUTES = 5
TOILET_WARNING_MINUTES = 14
TIME_LIMIT_FINE = 10
LATE_BACK_FINE = 50
DAILY_LIMIT_FINE = 100
ACTIVE_BREAK_ATTEMPT_FINE = 50

# ---------------- KEYBOARD ----------------
keyboard = [
    ["Start Work", "Off Work"],
    ["Smoke", "Toilet", "Eat"],
    ["Back to Seat"]
]
markup = ReplyKeyboardMarkup(keyboard, resize_keyboard=True)

# ---------------- DATA STORAGE ----------------
user_data = {}

# ---------------- HELPERS ----------------
def now_bd():
    return datetime.now(BANGLADESH_TZ)


def format_duration(td: timedelta):
    total_seconds = max(0, int(td.total_seconds()))
    hours = total_seconds // 3600
    minutes = (total_seconds % 3600) // 60
    seconds = total_seconds % 60
    return f"{hours:02d}:{minutes:02d}:{seconds:02d}"


def parse_duration(value):
    if isinstance(value, timedelta):
        return value
    if isinstance(value, (int, float)):
        return timedelta(seconds=float(value))
    return timedelta()


def dt_to_iso(dt):
    if not dt:
        return None
    return dt.astimezone(BANGLADESH_TZ).isoformat()


def iso_to_dt(value):
    if not value:
        return None
    try:
        return datetime.fromisoformat(value).astimezone(BANGLADESH_TZ)
    except Exception:
        return None


def time_str(dt: datetime):
    if not dt:
        return "N/A"
    return dt.astimezone(BANGLADESH_TZ).strftime("%I:%M:%S %p")


def datetime_str(dt: datetime):
    if not dt:
        return "N/A"
    return dt.astimezone(BANGLADESH_TZ).strftime("%Y-%m-%d %I:%M:%S %p")


def ordinal(n):
    if 10 <= n % 100 <= 20:
        return f"{n}th"
    return f"{n}{['th','st','nd','rd','th','th','th','th','th','th'][n % 10]}"


def is_admin(update: Update):
    user_id = update.effective_user.id if update.effective_user else None
    chat_id = update.effective_chat.id if update.effective_chat else None
    return user_id == ADMIN_CHAT_ID or chat_id == ADMIN_CHAT_ID


def period_bounds(reference=None):
    reference = (reference or now_bd()).astimezone(BANGLADESH_TZ)
    today_four = reference.replace(hour=4, minute=0, second=0, microsecond=0)
    if reference < today_four:
        start = today_four - timedelta(days=1)
    else:
        start = today_four
    return start, reference


def in_period(dt, start, end):
    if not dt:
        return False
    bd_dt = dt.astimezone(BANGLADESH_TZ)
    return start <= bd_dt <= end


def empty_user(chat_id=None, user_id=None, display_name=None):
    return {
        "Start Work": [],
        "Smoke": [],
        "Toilet": [],
        "Eat": [],
        "Back to Seat": [],
        "Off Work": None,
        "Away": None,
        "warning_sent": False,
        "chat_id": chat_id,
        "user_id": user_id,
        "display_name": display_name,
        "History": [],
        "Break History": [],
        "Warning History": [],
        "Fine History": []
    }


def ensure_user(username, chat_id=None, user_id=None, display_name=None):
    if username not in user_data:
        user_data[username] = empty_user(chat_id=chat_id, user_id=user_id, display_name=display_name)
    u = user_data[username]
    defaults = empty_user(chat_id=chat_id, user_id=user_id, display_name=display_name)
    for key, value in defaults.items():
        if key not in u:
            u[key] = value
    if chat_id is not None:
        u["chat_id"] = chat_id
    if user_id is not None:
        u["user_id"] = user_id
    if display_name is not None:
        u["display_name"] = display_name
    return u


def get_username(update: Update):
    user = update.effective_user
    if not user:
        return "unknown"
    return user.username or user.first_name or str(user.id)


def record_history(username, action, timestamp=None, details=None, status="✅", fine=0):
    u = ensure_user(username)
    timestamp = timestamp or now_bd()
    u["History"].append({
        "timestamp": dt_to_iso(timestamp),
        "action": action,
        "details": details or "",
        "status": status,
        "fine": fine
    })


def add_warning(username, warning_type, message, fine=0, timestamp=None):
    u = ensure_user(username)
    timestamp = timestamp or now_bd()
    warning = {
        "timestamp": dt_to_iso(timestamp),
        "type": warning_type,
        "message": message,
        "fine": fine,
        "status": "❌"
    }
    u["Warning History"].append(warning)
    if fine:
        u["Fine History"].append(warning.copy())
    record_history(username, warning_type, timestamp, message, "❌", fine)
    save_data()


def serialize_user(u):
    return {
        "Start Work": [dt_to_iso(dt) for dt in u.get("Start Work", [])],
        "Smoke": [parse_duration(td).total_seconds() for td in u.get("Smoke", [])],
        "Toilet": [parse_duration(td).total_seconds() for td in u.get("Toilet", [])],
        "Eat": [parse_duration(td).total_seconds() for td in u.get("Eat", [])],
        "Back to Seat": [
            {
                "action": row.get("action"),
                "time": dt_to_iso(row.get("time")),
                "duration": parse_duration(row.get("duration")).total_seconds()
            }
            for row in u.get("Back to Seat", [])
        ],
        "Off Work": dt_to_iso(u.get("Off Work")),
        "Away": {
            "action": u.get("Away", {}).get("action"),
            "time": dt_to_iso(u.get("Away", {}).get("time"))
        } if u.get("Away") else None,
        "warning_sent": bool(u.get("warning_sent", False)),
        "chat_id": u.get("chat_id"),
        "user_id": u.get("user_id"),
        "display_name": u.get("display_name"),
        "History": u.get("History", []),
        "Break History": u.get("Break History", []),
        "Warning History": u.get("Warning History", []),
        "Fine History": u.get("Fine History", [])
    }


def deserialize_user(raw):
    u = empty_user(
        chat_id=raw.get("chat_id"),
        user_id=raw.get("user_id"),
        display_name=raw.get("display_name")
    )
    u["Start Work"] = [dt for dt in (iso_to_dt(value) for value in raw.get("Start Work", [])) if dt]
    for act in ["Smoke", "Toilet", "Eat"]:
        u[act] = [parse_duration(value) for value in raw.get(act, [])]
    u["Back to Seat"] = []
    for row in raw.get("Back to Seat", []):
        u["Back to Seat"].append({
            "action": row.get("action"),
            "time": iso_to_dt(row.get("time")),
            "duration": parse_duration(row.get("duration"))
        })
    u["Off Work"] = iso_to_dt(raw.get("Off Work"))
    if raw.get("Away"):
        u["Away"] = {
            "action": raw["Away"].get("action"),
            "time": iso_to_dt(raw["Away"].get("time"))
        }
    u["warning_sent"] = bool(raw.get("warning_sent", False))
    u["History"] = raw.get("History", [])
    u["Break History"] = raw.get("Break History", [])
    u["Warning History"] = raw.get("Warning History", [])
    u["Fine History"] = raw.get("Fine History", [])
    return u


def save_data():
    os.makedirs(BASE_DIR, exist_ok=True)
    payload = {username: serialize_user(u) for username, u in user_data.items()}
    with open(DATA_FILE, "w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)


def load_data():
    global user_data
    if not os.path.exists(DATA_FILE):
        user_data = {}
        return
    try:
        with open(DATA_FILE, "r", encoding="utf-8") as f:
            payload = json.load(f)
        user_data = {username: deserialize_user(raw) for username, raw in payload.items()}
    except Exception as exc:
        print(f"Could not load {DATA_FILE}: {exc}")
        user_data = {}


def reset_current_data():
    for username in list(user_data.keys()):
        u = ensure_user(username)
        for key in ["Start Work", "Smoke", "Toilet", "Eat", "Back to Seat"]:
            u[key] = []
        u["Off Work"] = None
        u["Away"] = None
        u["warning_sent"] = False
    save_data()


def get_period_breaks(u, start, end):
    rows = []
    for row in u.get("Break History", []):
        ts = iso_to_dt(row.get("start_time"))
        if in_period(ts, start, end):
            rows.append(row)
    return rows


def get_period_warnings(u, start, end):
    rows = []
    for row in u.get("Warning History", []):
        ts = iso_to_dt(row.get("timestamp"))
        if in_period(ts, start, end):
            rows.append(row)
    return rows


def get_period_fines_total(u, start, end):
    total = 0
    for row in u.get("Fine History", []):
        ts = iso_to_dt(row.get("timestamp"))
        if in_period(ts, start, end):
            total += float(row.get("fine") or 0)
    return total


def calculate_current_totals(u, reference=None):
    reference = reference or now_bd()
    total_away = timedelta()
    for act in ["Smoke", "Toilet", "Eat"]:
        total_away += sum((parse_duration(td) for td in u.get(act, [])), timedelta())
    if u.get("Away") and u["Away"].get("time"):
        total_away += reference - u["Away"]["time"]
    if u.get("Start Work"):
        end_time = u.get("Off Work") or reference
        gross = end_time - u["Start Work"][0]
    else:
        gross = timedelta()
    net = gross - total_away
    if net.total_seconds() < 0:
        net = timedelta()
    return gross, total_away, net


def set_columns(ws, widths):
    for idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def style_sheet(ws):
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if cell.value == "❌":
                cell.fill = PatternFill("solid", fgColor="FFC7CE")
                cell.font = Font(color="9C0006", bold=True)
            elif cell.value == "✅":
                cell.fill = PatternFill("solid", fgColor="C6EFCE")
                cell.font = Font(color="006100", bold=True)


def generate_excel_report(start=None, end=None):
    if Workbook is None:
        raise RuntimeError("openpyxl is required to generate Excel reports. Install it with: pip install openpyxl")

    os.makedirs(REPORTS_DIR, exist_ok=True)
    end = end or now_bd()
    start = start or period_bounds(end)[0]
    filename = f"work_report_{start.strftime('%Y%m%d_%H%M%S')}_to_{end.strftime('%Y%m%d_%H%M%S')}.xlsx"
    path = os.path.join(REPORTS_DIR, filename)

    wb = Workbook()
    attendance = wb.active
    attendance.title = "Attendance"
    attendance.append([
        "Status", "User", "Telegram Chat ID", "Start Work", "Off Work",
        "Gross Work Time", "Total Away Time", "Net Working Time", "Total Fine"
    ])

    break_sheet = wb.create_sheet("Break History")
    break_sheet.append(["Status", "User", "Break Type", "Start Time", "End Time", "Duration", "Fine", "Notes"])

    warning_sheet = wb.create_sheet("Warnings & Fines")
    warning_sheet.append(["Status", "User", "Timestamp", "Type", "Warning", "Fine"])

    summary_sheet = wb.create_sheet("Summary")
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["Report Period Start", datetime_str(start)])
    summary_sheet.append(["Report Period End", datetime_str(end)])
    summary_sheet.append(["Timezone", "Asia/Dhaka"])
    summary_sheet.append(["Admin Chat ID", ADMIN_CHAT_ID])
    summary_sheet.append([])
    summary_sheet.append(["User", "Start Count", "Break Count", "Warning Count", "Total Away", "Net Working", "Total Fine"])

    total_users = 0
    total_fines = 0
    total_warning_count = 0

    for username, u in sorted(user_data.items()):
        total_users += 1
        start_times = [dt for dt in u.get("Start Work", []) if in_period(dt, start, end)]
        off_work = u.get("Off Work") if in_period(u.get("Off Work"), start, end) else None
        gross, total_away, net = calculate_current_totals(u, end)
        fine_total = get_period_fines_total(u, start, end)
        warnings = get_period_warnings(u, start, end)
        total_fines += fine_total
        total_warning_count += len(warnings)
        status = "❌" if warnings or fine_total else "✅"

        attendance.append([
            status,
            username,
            u.get("chat_id") or "",
            "\n".join(datetime_str(dt) for dt in start_times),
            datetime_str(off_work) if off_work else "",
            format_duration(gross),
            format_duration(total_away),
            format_duration(net),
            fine_total
        ])

        break_count = 0
        for row in get_period_breaks(u, start, end):
            break_count += 1
            break_sheet.append([
                row.get("status") or "✅",
                username,
                row.get("action", ""),
                datetime_str(iso_to_dt(row.get("start_time"))),
                datetime_str(iso_to_dt(row.get("end_time"))),
                format_duration(parse_duration(row.get("duration_seconds", 0))),
                row.get("fine", 0),
                row.get("notes", "")
            ])

        for row in warnings:
            warning_sheet.append([
                "❌",
                username,
                datetime_str(iso_to_dt(row.get("timestamp"))),
                row.get("type", ""),
                row.get("message", ""),
                row.get("fine", 0)
            ])

        summary_sheet.append([
            username,
            len(start_times),
            break_count,
            len(warnings),
            format_duration(total_away),
            format_duration(net),
            fine_total
        ])

    summary_sheet.append([])
    summary_sheet.append(["Total Users", total_users])
    summary_sheet.append(["Total Warnings", total_warning_count])
    summary_sheet.append(["Total Fines", total_fines])

    set_columns(attendance, [10, 22, 18, 30, 24, 18, 18, 18, 14])
    set_columns(break_sheet, [10, 22, 16, 24, 24, 16, 12, 42])
    set_columns(warning_sheet, [10, 22, 24, 24, 60, 12])
    set_columns(summary_sheet, [24, 28, 16, 16, 16, 18, 14])

    for ws in [attendance, break_sheet, warning_sheet, summary_sheet]:
        style_sheet(ws)
        ws.freeze_panes = "A2"

    wb.save(path)
    return path


async def send_excel_to_admin(context: ContextTypes.DEFAULT_TYPE, path, caption):
    with open(path, "rb") as report_file:
        await context.bot.send_document(
            chat_id=ADMIN_CHAT_ID,
            document=report_file,
            filename=os.path.basename(path),
            caption=caption
        )

# ---------------- REAL-TIME LIMIT CHECK ----------------
async def check_active_breaks(context: ContextTypes.DEFAULT_TYPE):
    now = now_bd()

    for user, u in user_data.items():
        active = u.get("Away")
        if not active or not active.get("time"):
            continue

        action = active["action"]
        start_time = active["time"]
        elapsed_minutes = (now - start_time).total_seconds() / 60

        if action == "Smoke" and elapsed_minutes >= SMOKE_WARNING_MINUTES and not u.get("warning_sent", False):
            await context.bot.send_message(
                chat_id=u["chat_id"],
                text=(
                    f"🚨 @{user}\n\n"
                    f"🚬 <b>Smoke Time Limit Exceeded</b>\n\n"
                    "⚠️ You have exceeded your smoke time limit.\n\n"
                    "👉 Please back to seat immediately or you will get <b>10$ fine</b>."
                ),
                parse_mode=ParseMode.HTML
            )
            u["warning_sent"] = True
            add_warning(user, "Smoke Time Limit Exceeded", "Smoke break exceeded 5 minutes", TIME_LIMIT_FINE, now)

        elif action == "Toilet" and elapsed_minutes >= TOILET_WARNING_MINUTES and not u.get("warning_sent", False):
            await context.bot.send_message(
                chat_id=u["chat_id"],
                text=(
                    f"🚨 @{user}\n\n"
                    f"🚻 <b>Toilet Time Limit Exceeded</b>\n\n"
                    "⚠️ You have exceeded your toilet time limit.\n\n"
                    "👉 Please back to seat immediately or you will get <b>10$ fine</b>."
                ),
                parse_mode=ParseMode.HTML
            )
            u["warning_sent"] = True
            add_warning(user, "Toilet Time Limit Exceeded", "Toilet break exceeded 14 minutes", TIME_LIMIT_FINE, now)

# ---------------- COMMANDS ----------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = get_username(update)
    ensure_user(
        user,
        chat_id=update.effective_chat.id,
        user_id=update.effective_user.id,
        display_name=update.effective_user.full_name
    )
    save_data()

    msg = (
        "✨ <b>Welcome to Work Tracker Bot</b>\n\n"
        "📌 <b>Status:</b> Ready to track your activity\n"
        "📱 Use the buttons below to control your workflow\n\n"
        "🚀 <b>Let’s get productive!</b>"
    )
    await update.message.reply_text(msg, reply_markup=markup, parse_mode=ParseMode.HTML)


async def reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("🚫 <b>Access Denied</b>\n\nOnly admin can use /reset.", parse_mode=ParseMode.HTML)
        return

    reset_current_data()
    await update.message.reply_text(
        "🔄 <b>System Reset Successful</b>\n\n"
        "🗑 Current user activity data has been cleared\n"
        "📚 Complete history, warning history, fines, and archived reports are preserved\n"
        "⚙️ Fresh tracking session is ready",
        parse_mode=ParseMode.HTML
    )


async def report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("🚫 <b>Access Denied</b>\n\nOnly admin can use /report.", parse_mode=ParseMode.HTML)
        return

    start_period, end_period = period_bounds(now_bd())
    try:
        path = generate_excel_report(start_period, end_period)
        with open(path, "rb") as report_file:
            await update.message.reply_document(
                document=report_file,
                filename=os.path.basename(path),
                caption=(
                    "📊 Work Tracker Excel Report\n"
                    f"Period: {datetime_str(start_period)} to {datetime_str(end_period)}\n"
                    "Archived in /reports folder."
                )
            )
    except Exception as exc:
        await update.message.reply_text(f"❌ <b>Report generation failed</b>\n\n<code>{exc}</code>", parse_mode=ParseMode.HTML)

# ---------------- BUTTON HANDLER ----------------
async def handle_button(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = get_username(update)
    action = update.message.text
    now = now_bd()

    u = ensure_user(user, chat_id=update.effective_chat.id, user_id=update.effective_user.id, display_name=update.effective_user.full_name)
    active_break = u["Away"]

    # ---------- PREVENT ACTION BEFORE START WORK ----------
    if action in ["Smoke", "Toilet", "Eat", "Off Work"] and not u["Start Work"]:
        await update.message.reply_text(
            "⚠️ <b>Action Not Allowed!</b>\n"
            "🚫 You cannot click <b>Smoke, Toilet, Eat</b> or <b>Off Work</b> before starting your work.\n"
            "🟢 Please click <b>Start Work</b> to officially begin your workday and enable other buttons.\n"
            "📝 Follow the workflow: Start Work → Breaks → Back to Seat → Off Work.\n"
            "💡 <b>অবশ্যই প্রথমে Start Work ক্লিক করুন। অন্য বাটনগুলো সক্রিয় হবে তার পরেই।</b>\n"
            "📌 প্রফেশনাল নির্দেশনা অনুযায়ী কার্যক্রম অনুসরণ করুন।",
            parse_mode=ParseMode.HTML
        )
        return

    # ---------- DAILY LIMIT CHECK ----------
    if action == "Smoke" and len(u["Smoke"]) >= SMOKE_LIMIT_COUNT:
        add_warning(user, "Daily Smoke Limit Exceeded", "User attempted more than 5 Smoke breaks in one day", DAILY_LIMIT_FINE, now)
        await update.message.reply_text(
            "🚫 <b>Daily Smoke Limit Exceeded</b>\n"
            "💰 You cannot take more than 5 Smoke breaks per day\n"
            "<b>Violation may result in $100 fine</b>",
            parse_mode=ParseMode.HTML
        )
        return

    if action == "Toilet" and len(u["Toilet"]) >= TOILET_LIMIT_COUNT:
        add_warning(user, "Daily Toilet Limit Exceeded", "User attempted more than 5 Toilet breaks in one day", DAILY_LIMIT_FINE, now)
        await update.message.reply_text(
            "🚫 <b>Daily Toilet Limit Exceeded</b>\n"
            "💰 You cannot take more than 5 Toilet breaks per day\n"
            "<b>Violation may result in $100 fine</b>",
            parse_mode=ParseMode.HTML
        )
        return

    if action == "Eat" and len(u["Eat"]) >= EAT_LIMIT_COUNT:
        add_warning(user, "Daily Eat Limit Exceeded", "User attempted more than 1 Eat break in one day", DAILY_LIMIT_FINE, now)
        await update.message.reply_text(
            "🚫 <b>Daily Eat Limit Exceeded</b>\n"
            "💰 You cannot take more than 1 Eat break per day\n"
            "<b>Violation may result in $100 fine</b>",
            parse_mode=ParseMode.HTML
        )
        return

    # ---------- VALIDATION ----------
    if action in ["Smoke", "Toilet", "Eat"]:
        if not u["Start Work"]:
            await update.message.reply_text("⚠️ <b>Work Not Started</b>\n👉 Please click <b>Start Work</b> before taking a break", parse_mode=ParseMode.HTML)
            return

        if active_break:
            add_warning(user, "Break Already Active", f"Attempted to start {action} while {active_break['action']} break was active", ACTIVE_BREAK_ATTEMPT_FINE, now)
            await update.message.reply_text(
                f"🚫 <b>BREAK ALREADY ACTIVE</b>\n\n"
                f"🕒 You are currently on <b>{active_break['action']}</b> break\n\n"
                "⚠️ <b>Action Required:</b>\n"
                "👉 You must click <b>Back to Seat</b> before starting any new activity\n\n"
                "💰 <b>Warning:</b> If you delay returning and do not click Back to Seat on time,\n"
                "a <b>$50 fine</b> may be applied\n\n"
                "🚨 <i>Please follow the process strictly to avoid penalties</i>",
                parse_mode=ParseMode.HTML
            )
            return

    # ---------- BACK TO SEAT ----------
    if action == "Back to Seat":
        if not active_break:
            await update.message.reply_text(
                "⚠️ <b>NO ACTIVE BREAK DETECTED</b>\n\n"
                "🚫 You are not currently on any break\n\n"
                "👉 Please start a valid activity first:\n"
                "<b>Smoke / Eat / Toilet</b>\n\n"
                "💡 After completing your break, always click <b>Back to Seat</b>\n"
                "💰 Late response may result in a <b>$50 fine</b>",
                parse_mode=ParseMode.HTML
            )
            return

        away_action = active_break["action"]
        start_time = active_break["time"]
        duration = now - start_time
        fine = 0
        status = "✅"
        notes = "Normal record"

        if away_action == "Smoke" and duration >= timedelta(minutes=SMOKE_WARNING_MINUTES):
            fine += LATE_BACK_FINE
            status = "❌"
            notes = "Late Back to Seat after Smoke time limit"
        elif away_action == "Toilet" and duration >= timedelta(minutes=TOILET_WARNING_MINUTES):
            fine += LATE_BACK_FINE
            status = "❌"
            notes = "Late Back to Seat after Toilet time limit"

        u["Back to Seat"].append({"action": away_action, "time": now, "duration": duration})
        u[away_action].append(duration)
        u["Break History"].append({
            "action": away_action,
            "start_time": dt_to_iso(start_time),
            "end_time": dt_to_iso(now),
            "duration_seconds": duration.total_seconds(),
            "status": status,
            "fine": fine,
            "notes": notes
        })
        record_history(user, f"{away_action} End", now, f"Duration: {format_duration(duration)}", status, fine)
        record_history(user, "Back to Seat", now, f"Returned from {away_action}", status, fine)
        if fine:
            add_warning(user, "Late Back to Seat", notes, fine, now)

        u["Away"] = None
        u["warning_sent"] = False
        save_data()

        total_duration = sum(u[away_action], timedelta())
        total_count = len(u[away_action])

        msg = (
            "🪑 <b>Back to Seat Recorded</b>\n\n"
            f"📌 <b>Break Type:</b> {away_action}\n"
            f"⌚ <b>Return Time:</b> {time_str(now)}\n\n"
            f"⏱ <b>Session Duration:</b> {format_duration(duration)}\n"
            f"📊 <b>Total Today:</b> {format_duration(total_duration)}\n"
            f"🔢 <b>Total Count:</b> {total_count}\n"
            f"💰 <b>Fine:</b> ${fine}\n\n"
            "⚠️ <i>Do not click Off Work if you are still working</i>"
        )
        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

    # ---------- START WORK ----------
    if action == "Start Work":
        if u["Start Work"] and not u["Off Work"]:
            await update.message.reply_text(
                "⚠️ <b>Work Already Started</b>\n\n"
                "📌 You haven't ended your previous session\n"
                "👉 Click <b>Off Work</b> before starting again",
                parse_mode=ParseMode.HTML
            )
            return

        u["Start Work"].append(now)
        u["Off Work"] = None
        record_history(user, "Start Work", now, "Work session started", "✅", 0)
        save_data()
        count = ordinal(len(u["Start Work"]))

        msg = (
            "🟢 <b>Work Session Started</b>\n\n"
            f"⌚ <b>Start Time:</b> {time_str(now)}\n"
            f"📊 <b>Session Count:</b> {count}\n\n"
            "💼 Stay focused and productive!"
        )
        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

    # ---------- OFF WORK ----------
    if action == "Off Work":
        if not u["Start Work"]:
            await update.message.reply_text("⚠️ <b>No Active Work Session</b>\n👉 Click <b>Start Work</b> first", parse_mode=ParseMode.HTML)
            return

        u["Off Work"] = now
        record_history(user, "Off Work", now, "Work session completed", "✅", 0)
        save_data()

        summary = ""
        total_away_time = timedelta()

        for act in ["Smoke", "Toilet", "Eat"]:
            count = len(u[act])
            duration = sum(u[act], timedelta())
            total_away_time += duration
            summary += f"• <b>{act}</b> → Count: {count}, Time: {format_duration(duration)}\n"

        start_times = ", ".join([time_str(t) for t in u["Start Work"]])
        total_working = (now - u["Start Work"][0]) - total_away_time
        if total_working.total_seconds() < 0:
            total_working = timedelta()

        msg = (
            "📊 <b>Work Summary Report</b>\n\n"
            "✅ <b>Status:</b> Work Completed\n\n"
            f"🕒 <b>Start Times:</b> {start_times}\n\n"
            f"{summary}\n"
            f"⏱ <b>Total Away Time:</b> {format_duration(total_away_time)}\n"
            f"⏱ <b>Net Working Time:</b> {format_duration(total_working)}\n\n"
            "🚀 <i>Click Start Work to begin a new session</i>"
        )

        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

    # ---------- BREAKS ----------
    if action == "Eat":
        u["Away"] = {"action": action, "time": now}
        u["warning_sent"] = False
        record_history(user, "Eat Start", now, "Eat break started", "✅", 0)
        save_data()
        today_count = len(u[action])
        start_work_time = time_str(u["Start Work"][0])

        msg = (
            "🍽 <b>Break Started: Eat</b>\n\n"
            f"⌚ <b>Start Time:</b> {time_str(now)}\n"
            f"📊 <b>Today's Count:</b> {today_count}\n"
            f"🕒 <b>Work Start:</b> {start_work_time}\n\n"
            "⚠️ <b>Important Notice:</b>\n"
            "👉 You must click <b>Back to Seat</b> immediately after returning\n"
            "⏳ Any delay in response will be strictly monitored\n"
            "💰 <b>Penalty:</b> Late Back to Seat action will result in a <b>$50 fine</b>\n"
            "🚨 <i>Please ensure compliance</i>"
        )
        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

    if action == "Toilet":
        u["Away"] = {"action": action, "time": now}
        u["warning_sent"] = False
        record_history(user, "Toilet Start", now, "Toilet break started", "✅", 0)
        save_data()
        today_count = len(u[action])

        msg = (
            "🚻 <b>Break Started: Toilet</b>\n\n"
            f"⌚ <b>Start Time:</b> {time_str(now)}\n"
            f"📊 <b>Today's Count:</b> {today_count}\n\n"
            "⚠️ <b>Mandatory Action:</b>\n"
            "👉 Click <b>Back to Seat</b> immediately after returning\n"
            "⏳ Delay will be considered a violation\n"
            "💰 <b>$50 fine</b> will be applied for late action\n"
            "🚨 <i>Strict monitoring in place</i>"
        )
        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

    if action == "Smoke":
        u["Away"] = {"action": action, "time": now}
        u["warning_sent"] = False
        record_history(user, "Smoke Start", now, "Smoke break started", "✅", 0)
        save_data()
        today_count = len(u[action])

        msg = (
            "🚬 <b>Break Started: Smoke</b>\n\n"
            f"⌚ <b>Start Time:</b> {time_str(now)}\n"
            f"📊 <b>Today's Count:</b> {today_count}\n\n"
            "⚠️ <b>Important Instruction:</b>\n"
            "👉 You must click <b>Back to Seat</b> as soon as you return\n"
            "⏳ Any delay will be tracked\n"
            "💰 Late update will result in a <b>$50 fine</b>\n"
            "🚨 <i>Avoid penalties by responding on time</i>"
        )
        await update.message.reply_text(msg, parse_mode=ParseMode.HTML)
        return

# ---------------- SCHEDULED JOBS ----------------
async def scheduled_report(context: ContextTypes.DEFAULT_TYPE):
    end_period = now_bd()
    start_period = period_bounds(end_period)[0]
    try:
        path = generate_excel_report(start_period, end_period)
        await send_excel_to_admin(
            context,
            path,
            (
                "📊 Daily Work Tracker Excel Report\n"
                f"Period: {datetime_str(start_period)} to {datetime_str(end_period)}\n"
                "Archived in /reports folder."
            )
        )
    except Exception as exc:
        await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text=f"❌ Daily report generation failed: {exc}")


async def scheduled_reset(context: ContextTypes.DEFAULT_TYPE):
    reset_current_data()
    await context.bot.send_message(chat_id=ADMIN_CHAT_ID, text="🔄 Daily automatic reset completed at 4:00 AM Asia/Dhaka.")

# ---------------- BOT RUN ----------------
def main():
    if not TOKEN:
        raise RuntimeError("TOKEN environment variable is missing. Set TOKEN before running the bot.")

    load_data()

    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("report", report))
    app.add_handler(CommandHandler("reset", reset))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_button))

    app.job_queue.run_repeating(check_active_breaks, interval=30, first=30)
    app.job_queue.run_daily(
        scheduled_report,
        time=datetime_time(hour=3, minute=59, second=50, tzinfo=BANGLADESH_TZ),
        name="daily_excel_report"
    )
    app.job_queue.run_daily(
        scheduled_reset,
        time=datetime_time(hour=4, minute=0, second=0, tzinfo=BANGLADESH_TZ),
        name="daily_auto_reset"
    )

    print("Bot is running... ✅")
    app.run_polling()


if __name__ == "__main__":
    main()
